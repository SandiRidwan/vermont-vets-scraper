"""
Vermont Vets - VVMA Scraper v4.0
=================================
Strategi: Biarkan Angular app yang navigasi sendiri ke halaman berikut.
Kita intercept semua response dari service-router via Playwright route handler
(bukan JS fetch — itu cross-origin CORS blocked).

Install:
    pip install playwright pandas beautifulsoup4 openpyxl
    playwright install chromium

Run:
    python vvma_scraper_v4.py
"""

import asyncio
import re
import json
import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, Route, Request
from openpyxl.styles import Font, PatternFill

# ── CONFIG ────────────────────────────────────────────────────────────────────
TARGET_URL    = "https://www.vtvets.org/find-a-vet-search"
GMAPS_CSV     = "vermont_vets_comprehensive.csv"
OUTPUT_VVMA   = "vermont_vets_VVMA.csv"
OUTPUT_MERGED = "vermont_vets_merged_final.csv"
OUTPUT_XLSX   = "vermont_vets_merged_final.xlsx"

# ── HELPERS ───────────────────────────────────────────────────────────────────
def ev(html):
    if not html: return ""
    soup = BeautifulSoup(html, "html.parser")
    for s in soup.find_all("strong"): s.decompose()
    return soup.get_text(strip=True)

def format_phone(raw: str) -> str:
    digits = re.sub(r'\D', '', str(raw))
    if len(digits) == 11 and digits.startswith('1'): digits = digits[1:]
    if len(digits) == 10:
        return f"+1 ({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return raw.strip()

def parse_name(title: str) -> tuple:
    name = re.sub(r"^(Dr\.|Dr)\s*", "", title.strip())
    name = re.sub(r"\b(DVM|VMD|PhD|MS|DACVIM|DACVS|DACVD|Jr\.?|Sr\.?|III|II)\b", "", name, flags=re.I)
    name = name.strip().rstrip(",").strip()
    parts = name.split()
    return (parts[0] if parts else ""), (" ".join(parts[1:]) if len(parts) > 1 else "")

def clean_url(u: str) -> str:
    if not u or not str(u).strip(): return ""
    u = str(u).strip().split("?")[0].rstrip("/")
    if not re.match(r'https?://|www\.', u): return ""
    if not u.startswith("http"): u = "http://" + u
    return u

def parse_record(r: dict) -> dict:
    first, last = parse_name(r.get("title", ""))
    top = r.get("top", []); left = r.get("left", []); right = r.get("right", [])
    addr1   = ev(top[0]["html"])   if len(top)   > 0 else ""
    city    = ev(left[0]["html"])  if len(left)  > 0 else ""
    state   = ev(left[1]["html"])  if len(left)  > 1 else ""
    zipcode = ev(left[2]["html"])  if len(left)  > 2 else ""
    phone   = ev(right[0]["html"]) if len(right) > 0 else ""
    clinic  = ev(right[1]["html"]) if len(right) > 1 else ""
    website = ev(right[2]["html"]) if len(right) > 2 else ""
    phone   = format_phone(phone) if re.sub(r'\D', '', phone) else ""
    website = clean_url(website)
    address = ", ".join(p for p in [addr1, city, state, zipcode] if p)
    return {
        "clinic_name": clinic, "first_name": first, "last_name": last,
        "email": "", "phone": phone, "address": address,
        "website": website, "source_query": "VVMA_directory",
    }

# ── PLAYWRIGHT SCRAPER (v4.1 - FIXED) ────────────────────────────────────────
async def scrape_vvma() -> pd.DataFrame:
    print("=" * 65)
    print("  VVMA Scraper v4.1 — FIXED Response Interceptor")
    print("=" * 65)

    all_records   = []
    pages_captured = set()
    total_pages   = {"value": 24}

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--disable-blink-features=AutomationControlled"])
        context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        page = await context.new_page()

        async def on_response(response):
            if "json" not in response.headers.get("content-type", ""):
                return

            url = response.url
            try:
                body = await response.json()

                # Tangkap SEMUA response yang punya "results"
                if isinstance(body, dict) and "results" in body:
                    results = body.get("results", [])
                    if not results:
                        return

                    # Deteksi nomor halaman
                    page_num = 1
                    if "pageNumber" in url:
                        m = re.search(r'pageNumber=(\d+)', url)
                        if m: page_num = int(m.group(1))
                    elif "current_page" in body:
                        page_num = body.get("current_page", 1)
                    elif "page" in body:
                        page_num = body.get("page", 1)

                    if page_num not in pages_captured:
                        pages_captured.add(page_num)
                        for r in results:
                            all_records.append(parse_record(r))
                        
                        print(f"  [Page {page_num:2d}] +{len(results)} records | Total: {len(all_records)}")

                        if len(pages_captured) >= total_pages["value"]:
                            print("  ✅ Semua halaman sudah terkumpul!")
            except:
                pass  # skip error

        page.on("response", on_response)

        # Step 1 + 2
        print(f"\nStep 1: Loading {TARGET_URL}...")
        await page.goto(TARGET_URL, wait_until="networkidle", timeout=60000)
        await asyncio.sleep(4)

        print("Step 2: Triggering search...")
        for selector in ['button[type="submit"]', 'button:has-text("Search")', 'button.mc-button', 'button']:
            try:
                btn = page.locator(selector).first
                if await btn.is_visible(timeout=3000):
                    await btn.click()
                    print(f"  ✓ Clicked: {selector}")
                    break
            except:
                continue
        await asyncio.sleep(6)

        # Step 3: Klik Next (ditingkatkan)
        print("\nStep 3: Klik Next sampai akhir...")
        for i in range(40):   # maksimal 40 kali
            try:
                # Beberapa kemungkinan tombol Next
                next_btn = page.locator("button:has-text('Next'), a:has-text('Next'), button[aria-label*='Next'], .pagination-next button, button mat-icon")
                if await next_btn.count() > 0:
                    btn = next_btn.first
                    if await btn.is_visible() and await btn.is_enabled():
                        await btn.click()
                        print(f"  Klik Next → halaman {i+2}")
                        await asyncio.sleep(3.5)   # lebih lama agar response sempat masuk
                    else:
                        break
                else:
                    break
            except:
                break

        await asyncio.sleep(8)   # tunggu response terakhir
        await browser.close()

    # Final processing
    df = pd.DataFrame(all_records)
    if df.empty:
        print("❌ Tidak ada data berhasil dikumpulkan.")
        return df

    df = df[df["first_name"] != ""].copy()
    df = df.drop_duplicates(subset=["first_name", "last_name"])
    df = df.sort_values(["last_name", "first_name"]).reset_index(drop=True)

    df.to_csv(OUTPUT_VVMA, index=False, encoding="utf-8-sig")
    print(f"\n  ✓ VVMA saved: {OUTPUT_VVMA}  ({len(df)} records)")
    return df

# ── GMAPS CLEANER ─────────────────────────────────────────────────────────────
def clean_gmaps(path: str) -> pd.DataFrame:
    print(f"\nCleaning Google Maps CSV...")
    df = pd.read_csv(path, encoding="utf-8-sig")
    df["address"] = (df["address"]
        .str.replace(r"^\s+", "", regex=True)
        .str.replace(r"\s+", " ", regex=True).str.strip())
    df["phone"] = df["phone"].apply(
        lambda x: format_phone(str(x)) if pd.notna(x) and str(x).strip() else "")
    df["website"] = df["website"].apply(
        lambda u: str(u).strip().split("?")[0].rstrip("/") if pd.notna(u) and str(u).strip() else "")

    def fix_names(row):
        cn = str(row.get("clinic_name", ""))
        if re.search(r'\bDVM\b|\bDr\.?\b', cn, re.I):
            name = re.sub(r'\bDVM\b|\bVMD\b', '', cn, flags=re.I)
            name = re.sub(r'^Dr\.?\s*', '', name, flags=re.I).strip().rstrip(",").strip()
            parts = name.split()
            return pd.Series({"first_name": parts[0] if parts else "",
                               "last_name":  " ".join(parts[1:]) if len(parts)>1 else ""})
        return pd.Series({"first_name": "", "last_name": ""})

    fixed = df.apply(fix_names, axis=1)
    df["first_name"] = fixed["first_name"]
    df["last_name"]  = fixed["last_name"]
    df = df[df["address"].str.contains(r'\bVT\b|\bVermont\b', na=False, flags=re.I)]
    skip = ["PetVet Vaccination","Vetco",r"\bPetco\b","PetSmart","Dollar","Tractor Supply"]
    df = df[~df["clinic_name"].str.contains("|".join(skip), case=False, na=False, regex=True)]
    df["source_query"] = "Google_Maps_" + df["source_query"].fillna("")
    print(f"  ✓ GMaps records: {len(df)}")
    return df.reset_index(drop=True)

# ── MERGER ────────────────────────────────────────────────────────────────────
def merge_data(df_vvma: pd.DataFrame, df_gmaps: pd.DataFrame) -> pd.DataFrame:
    print("\nMerging VVMA + Google Maps...")
    COLS = ["clinic_name","first_name","last_name","email",
            "phone","address","website","source_query"]
    df_all = pd.concat([df_vvma[COLS], df_gmaps[COLS]], ignore_index=True)
    df_all["_score"] = (
        (df_all["email"]   != "").astype(int) * 3 +
        (df_all["phone"]   != "").astype(int) * 2 +
        (df_all["website"] != "").astype(int) * 1)
    df_all["_pd"] = df_all["phone"].apply(lambda x: re.sub(r'\D','',str(x)))
    df_all = df_all.sort_values("_score", ascending=False)
    has_p = df_all[df_all["_pd"].str.len()==10].drop_duplicates(subset=["_pd"], keep="first")
    no_p  = df_all[df_all["_pd"].str.len()!=10].copy()
    no_p["_ck"] = no_p["clinic_name"].str.lower().str.replace(r'[^a-z0-9]','',regex=True).str[:20]
    no_p  = no_p.drop_duplicates(subset=["_ck"], keep="first")
    df_m  = pd.concat([has_p, no_p], ignore_index=True)
    df_m  = df_m.drop(columns=["_score","_pd","_ck"], errors="ignore")
    df_m  = df_m.sort_values(["last_name","first_name","clinic_name"]).reset_index(drop=True)
    print(f"  ✓ Merged: {len(df_m)} records  "
          f"(VVMA: {(df_m['source_query']=='VVMA_directory').sum()} | "
          f"GMaps: {df_m['source_query'].str.startswith('Google_Maps').sum()})")
    return df_m

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def export_excel(df: pd.DataFrame, path: str):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vermont Vets")
        ws = writer.sheets["Vermont Vets"]
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w+4, 55)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
    print(f"  ✓ Excel: {path}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
async def main():
    df_vvma = await scrape_vvma()
    if df_vvma.empty:
        print("❌ VVMA scraping gagal.")
        return

    try:
        df_gmaps = clean_gmaps(GMAPS_CSV)
    except FileNotFoundError:
        print(f"  ⚠ {GMAPS_CSV} tidak ditemukan — VVMA only")
        df_gmaps = pd.DataFrame(columns=["clinic_name","first_name","last_name",
                                          "email","phone","address","website","source_query"])

    df_final = merge_data(df_vvma, df_gmaps)
    df_final.to_csv(OUTPUT_MERGED, index=False, encoding="utf-8-sig")
    print(f"  ✓ CSV: {OUTPUT_MERGED}")
    export_excel(df_final, OUTPUT_XLSX)

    print(f"\n{'='*65}")
    print(f"  ✨ DONE!  Total: {len(df_final)} records")
    print(f"  📧 Email  : {(df_final['email']!='').sum()}")
    print(f"  📞 Phone  : {(df_final['phone']!='').sum()}")
    print(f"  🌐 Website: {(df_final['website']!='').sum()}")
    print(f"{'='*65}")
    print(df_final[["first_name","last_name","phone","clinic_name"]].head(8).to_string(index=False))

if __name__ == "__main__":
    asyncio.run(main())