import asyncio
import re
import time
import pandas as pd
from playwright.async_api import async_playwright
from playwright_stealth import Stealth

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
QUERIES = [
    # Addison County
    "veterinarian Middlebury VT", "veterinarian Bristol VT", "veterinarian Vergennes VT",
    # Bennington County
    "veterinarian Bennington VT", "veterinarian Manchester VT", "veterinarian Dorset VT",
    # Caledonia County
    "veterinarian St. Johnsbury VT", "veterinarian Lyndon VT", "veterinarian Hardwick VT",
    # Chittenden County
    "veterinarian Burlington VT", "veterinarian South Burlington VT", "veterinarian Essex VT",
    "veterinarian Colchester VT", "veterinarian Shelburne VT", "veterinarian Williston VT", "veterinarian Milton VT",
    # Franklin & Grand Isle
    "veterinarian St. Albans VT", "veterinarian Swanton VT", "veterinarian South Hero VT",
    # Lamoille County
    "veterinarian Stowe VT", "veterinarian Morrisville VT", "veterinarian Cambridge VT",
    # Orange & Orleans
    "veterinarian Randolph VT", "veterinarian Bradford VT", "veterinarian Newport VT", "veterinarian Derby VT",
    # Rutland County
    "veterinarian Rutland VT", "veterinarian Brandon VT", "veterinarian Poultney VT",
    # Washington County
    "veterinarian Montpelier VT", "veterinarian Barre VT", "veterinarian Waterbury VT",
    # Windham County
    "veterinarian Brattleboro VT", "veterinarian Wilmington VT", "veterinarian Rockingham VT",
    # Windsor County
    "veterinarian Hartford VT", "veterinarian White River Junction VT", "veterinarian Springfield VT",
    "veterinarian Woodstock VT", "veterinarian Ludlow VT",
    # Catch-all
    "animal hospital Vermont", "emergency vet Vermont", "equine veterinarian Vermont", "livestock vet Vermont"
]

OUTPUT_CSV  = "vermont_vets_comprehensive.csv"
OUTPUT_XLSX = "vermont_vets_comprehensive.xlsx"
BACKUP_CSV  = "vermont_vets_BACKUP.csv"

SCROLL_PAUSE = 2.5
DETAIL_PAUSE = 2.5
EMAIL_PAUSE  = 1.5
MAX_RETRIES  = 2        # retry per listing kalau gagal
QUERY_DELAY  = 4        # jeda antar query (detik)
BACKUP_EVERY = 5        # backup setiap N query

# ── HELPERS ───────────────────────────────────────────────────────────────────
EMAIL_RE   = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
SKIP_EMAIL = ["@sentry","@example","noreply",".png",".jpg","wixpress",
              "squarespace","@2x","support@","schema.org","@siteground",
              "your@","email@","info@info","test@"]

def clean_email(html: str) -> str:
    for m in EMAIL_RE.findall(html):
        if not any(s in m.lower() for s in SKIP_EMAIL):
            return m.lower()
    return ""

def format_phone(raw: str) -> str:
    """Normalisasi phone ke format +1 (XXX) XXX-XXXX."""
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f"+1 ({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return raw.strip()

def split_name(clinic_name: str) -> tuple[str, str]:
    """
    Best-effort split nama dokter dari nama klinik.
    Contoh: 'Green Mountain Animal Hospital' → ('Green', 'Hospital')
    Nama dokter biasanya ada di About section Maps, bukan di nama klinik.
    Ini placeholder — enrichment manual tetap disarankan untuk first/last name.
    """
    parts = clinic_name.split()
    return (parts[0] if parts else ""), (parts[-1] if len(parts) > 1 else "")

async def dismiss_consent(page):
    for sel in ['button:has-text("Accept all")', 'button:has-text("Agree")',
                'button:has-text("Setuju")', 'button:has-text("Terima semua")',
                '[aria-label="Accept all"]', 'form button:first-of-type']:
        try:
            await page.click(sel, timeout=2000)
            await asyncio.sleep(1)
            return
        except:
            pass

async def scroll_to_end(page):
    feed_sel = 'div[role="feed"]'
    try:
        await page.wait_for_selector(feed_sel, timeout=8000)
        feed = page.locator(feed_sel).first
        prev_count = 0
        stale = 0
        for _ in range(60):
            await feed.evaluate("el => el.scrollTop = el.scrollHeight")
            await asyncio.sleep(SCROLL_PAUSE)
            html = await page.content()
            if "You've reached the end" in html or "No more results" in html:
                break
            count = await page.locator('a[href*="/maps/place/"]').count()
            if count == prev_count:
                stale += 1
                if stale >= 3:
                    break
            else:
                stale = 0
                prev_count = count
    except:
        await page.keyboard.press("End")
        await asyncio.sleep(SCROLL_PAUSE)

async def extract_detail(page) -> dict:
    d = {"name": "", "phone": "", "address": "", "website": ""}
    try:
        d["name"] = await page.locator('h1.DUwDvf').first.inner_text(timeout=5000)
    except:
        pass

    # Phone — regex dari raw HTML (paling reliable lintas locale)
    try:
        html = await page.content()
        m = re.search(r'\+1\s?[\(]?\d{3}[\)\-\.\s]\s?[\d\s]{3}[\-\.\s]\d{4}', html)
        if m:
            d["phone"] = format_phone(m.group(0))
    except:
        pass

    try:
        d["address"] = await page.locator('[data-item-id="address"]').first.inner_text(timeout=2000)
    except:
        pass

    try:
        href = await page.locator('a[data-item-id="authority"]').first.get_attribute("href", timeout=2000)
        if href and "google.com" not in href:
            d["website"] = href
    except:
        pass

    return d

async def fetch_email_from_web(browser, website: str) -> str:
    if not website:
        return ""
    page = await browser.new_page()
    try:
        await page.goto(website, timeout=12000, wait_until="domcontentloaded")
        email = clean_email(await page.content())
        if not email:
            for sub in ["/contact", "/about", "/contact-us", "/team"]:
                try:
                    await page.goto(website.rstrip("/") + sub, timeout=8000,
                                    wait_until="domcontentloaded")
                    email = clean_email(await page.content())
                    if email:
                        break
                except:
                    continue
        await asyncio.sleep(EMAIL_PAUSE)
        return email
    except:
        return ""
    finally:
        await page.close()

# ── CORE SCRAPER ──────────────────────────────────────────────────────────────
async def scrape_query(browser, query: str, seen_ids: set) -> list:
    page = await browser.new_page()
    await page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
    records = []

    try:
        search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}?hl=en"
        await page.goto(search_url, wait_until="domcontentloaded")
        await asyncio.sleep(2)
        await dismiss_consent(page)
        await scroll_to_end(page)

        links = await page.locator('a[href*="/maps/place/"]').evaluate_all(
            "nodes => nodes.map(n => n.href)"
        )
        unique_links = list(set(links))
        print(f"   ✅ {len(unique_links)} listings found")

        for link in unique_links:
            m = re.search(r'place/([^/]+)/', link)
            place_id = m.group(1) if m else link
            if place_id in seen_ids:
                continue
            seen_ids.add(place_id)

            # ── Retry logic ──────────────────────────────────────────────────
            data = None
            for attempt in range(MAX_RETRIES):
                try:
                    await page.goto(link + "&hl=en", wait_until="domcontentloaded",
                                    timeout=20000)
                    await asyncio.sleep(DETAIL_PAUSE)
                    data = await extract_detail(page)
                    if data["name"]:
                        break
                except Exception as e:
                    if attempt == MAX_RETRIES - 1:
                        print(f"      ⚠ Skip (retry {MAX_RETRIES}x failed): {link[:60]}")
                    await asyncio.sleep(2)

            if not data or not data["name"]:
                continue

            dedup_key = f"{data['name']}|{data['address']}"
            if dedup_key in seen_ids:
                continue
            seen_ids.add(dedup_key)

            email = await fetch_email_from_web(browser, data["website"])
            first, last = split_name(data["name"])

            row = {
                "clinic_name":  data["name"],
                "first_name":   first,
                "last_name":    last,
                "email":        email,
                "phone":        data["phone"],
                "address":      data["address"],
                "website":      data["website"],
                "source_query": query,
            }
            records.append(row)
            print(f"      ✓ {data['name'][:28]:28} | {data['phone']:18} | {email or '—'}")

    except Exception as e:
        print(f"   ❌ Error: {e}")
    finally:
        await page.close()

    return records

# ── EXPORT ────────────────────────────────────────────────────────────────────
def export(df: pd.DataFrame, label: str = ""):
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vermont Vets")
        ws = writer.sheets["Vermont Vets"]
        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        # Header bold
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
    if label:
        print(f"   💾 Backup saved ({label}): {len(df)} records")

# ── MAIN ──────────────────────────────────────────────────────────────────────
async def main():
    all_data: list = []
    seen_ids: set  = set()
    start_time     = time.time()

    print("=" * 60)
    print("  VERMONT VETERINARIANS SCRAPER  v2.0")
    print(f"  {len(QUERIES)} queries | Target: 400+ records")
    print("=" * 60)

    async with Stealth().use_async(async_playwright()) as p:
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled",
                  "--no-sandbox", "--lang=en-US"]
        )

        for i, q in enumerate(QUERIES, 1):
            print(f"\n[{i:02d}/{len(QUERIES)}] {q}")
            results = await scrape_query(browser, q, seen_ids)
            all_data.extend(results)
            print(f"   📊 Running total: {len(all_data)} records")

            # Backup otomatis setiap BACKUP_EVERY query
            if i % BACKUP_EVERY == 0 and all_data:
                df_bak = pd.DataFrame(all_data).drop_duplicates(
                    subset=["clinic_name", "address"]
                )
                df_bak.to_csv(BACKUP_CSV, index=False, encoding="utf-8-sig")
                print(f"   💾 Auto-backup: {len(df_bak)} records → {BACKUP_CSV}")

            await asyncio.sleep(QUERY_DELAY)

        await browser.close()

    # ── Final export ──────────────────────────────────────────────────────────
    if not all_data:
        print("\n⚠ No data collected.")
        return

    df = pd.DataFrame(all_data)
    df.drop_duplicates(subset=["clinic_name", "address"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    export(df, label="final")

    elapsed = time.time() - start_time
    mins, secs = divmod(int(elapsed), 60)

    print("\n" + "=" * 60)
    print("  ✨ DONE!")
    print(f"  📋 Total unique records : {len(df)}")
    print(f"  📧 Records with email   : {df['email'].astype(bool).sum()} "
          f"({df['email'].astype(bool).mean()*100:.0f}%)")
    print(f"  📞 Records with phone   : {df['phone'].astype(bool).sum()} "
          f"({df['phone'].astype(bool).mean()*100:.0f}%)")
    print(f"  ⏱ Total runtime        : {mins}m {secs}s")
    print(f"  📂 Output: {OUTPUT_XLSX} & {OUTPUT_CSV}")
    print("=" * 60)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n⛔ Stopped by user. Progress was auto-saved to backup CSV.")
